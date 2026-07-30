"""Importador de tablas exportadas de ETABS (CSV, TSV, TXT y XLSX).

Es el camino que sirve cuando ETABS no está instalado en la máquina: el usuario
exporta desde ETABS (*File ▸ Export ▸ Tables*) y aquí se leen. Se reconocen por
sus columnas, no por su nombre, así que funcionan igual las tablas en español o
con encabezados reordenados:

* ``Element Forces - Beams`` / ``- Columns`` / ``- Frames`` → momentos, cortantes,
  axial y torsión por estación.
* ``Joint Displacements`` → desplazamientos y rotaciones (U1/U2/U3, R1/R2/R3 o
  Ux/Uy/Uz, Rx/Ry/Rz).
* ``Story Drifts`` / ``Diaphragm ... Drifts`` → derivas de entrepiso.
* ``Base Reactions`` → reacciones globales.

Un archivo puede traer varias tablas concatenadas: cada bloque ``TABLE: "…"`` se
lee por separado y todo se acumula en un único :class:`ResultSet`.
"""

from __future__ import annotations

import csv
import io
import os
import re
from typing import Dict, List, Optional, Sequence, Tuple

from .model import (
    DriftRow, FrameForceRow, JointDisplRow, ReactionRow, ResultSet,
)
from .units import UnitSystem, parse_units_label

_TABLE_TAG = re.compile(r'^\s*TABLE\s*:\s*"?([^"]*)"?', re.IGNORECASE)
_NUMERIC = re.compile(r"^[-+]?[0-9]*\.?[0-9]+(?:[eE][-+]?[0-9]+)?$")

#: Sinónimos de columna → nombre canónico interno.
_ALIASES: Dict[str, str] = {
    "story": "story", "nivel": "story", "storey": "story",
    "beam": "element", "column": "element", "frame": "element",
    "brace": "element", "barra": "element", "trabe": "element",
    "uniquename": "uniquename", "nombreunico": "uniquename",
    "label": "label", "etiqueta": "label",
    "outputcase": "case", "case": "case", "caso": "case",
    "casocarga": "case", "combo": "case", "combocase": "case",
    "casetype": "casetype", "steptype": "steptype", "tipopaso": "steptype",
    "stepnumber": "stepnum", "stepnum": "stepnum",
    "station": "station", "estacion": "station", "loc": "station",
    "p": "P", "axial": "P",
    "v2": "V2", "v3": "V3", "t": "T", "m2": "M2", "m3": "M3",
    "joint": "point", "point": "point", "nudo": "point", "pointobj": "point",
    "u1": "U1", "ux": "U1", "u2": "U2", "uy": "U2", "u3": "U3", "uz": "U3",
    "r1": "R1", "rx": "R1", "r2": "R2", "ry": "R2", "r3": "R3", "rz": "R3",
    "drift": "drift", "deriva": "drift",
    "direction": "direction", "direccion": "direction",
    "x": "x", "y": "y", "z": "z",
    "globalfx": "FX", "fx": "FX", "globalfy": "FY", "fy": "FY",
    "globalfz": "FZ", "fz": "FZ",
    "globalmx": "MX", "mx": "MX", "globalmy": "MY", "my": "MY",
    "globalmz": "MZ", "mz": "MZ",
}


def _norm(name: object) -> str:
    """Normaliza un encabezado: minúsculas y sin separadores ni acentos."""
    text = str(name or "").strip().lower()
    for accented, plain in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"),
                            ("ú", "u"), ("ñ", "n")):
        text = text.replace(accented, plain)
    return re.sub(r"[^a-z0-9]", "", text)


def _num(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text or not _NUMERIC.match(text):
        return None
    return float(text)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


# ── lectura del archivo a matriz de celdas ──────────────────────────────────


def read_cells(path: str) -> List[List[object]]:
    """Lee CSV/TSV/TXT/XLSX y devuelve la matriz de celdas cruda."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    return _read_delimited(path)


def _read_delimited(path: str) -> List[List[object]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                text = handle.read()
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover
        raise ValueError(f"No se pudo decodificar el archivo: {path}")

    sample = "\n".join(text.splitlines()[:40])
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    if delimiter == "," and sample.count(";") > sample.count(","):
        delimiter = ";"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def _read_xlsx(path: str) -> List[List[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depende del entorno
        raise ImportError(
            "Para leer .xlsx se necesita openpyxl (pip install openpyxl). "
            "También puedes exportar la tabla como CSV desde ETABS.") from exc
    book = load_workbook(path, read_only=True, data_only=True)
    cells: List[List[object]] = []
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells.append(list(row))
        cells.append([])          # separador entre hojas
    book.close()
    return cells


# ── interpretación ──────────────────────────────────────────────────────────


def _split_blocks(cells: Sequence[Sequence[object]]
                  ) -> List[Tuple[Optional[str], List[Sequence[object]]]]:
    """Parte la matriz en bloques, uno por cada ``TABLE: "…"``."""
    blocks: List[Tuple[Optional[str], List[Sequence[object]]]] = []
    current_name: Optional[str] = None
    current: List[Sequence[object]] = []
    for row in cells:
        first = _text(row[0]) if row else ""
        tag = _TABLE_TAG.match(first) if first else None
        if tag:
            if current:
                blocks.append((current_name, current))
            current_name = tag.group(1).strip() or None
            current = []
            continue
        current.append(row)
    if current:
        blocks.append((current_name, current))
    return blocks


def _find_header(rows: Sequence[Sequence[object]]) -> Tuple[int, List[str]]:
    """Localiza el renglón de encabezados y devuelve sus nombres canónicos."""
    best: Tuple[int, List[str]] = (-1, [])
    for index, row in enumerate(rows[:30]):
        canonical = [_ALIASES.get(_norm(c), "") for c in row]
        named = {c for c in canonical if c}
        has_id = bool({"story", "element", "point", "case"} & named)
        has_data = bool({"station", "M3", "V2", "U3", "U1", "drift", "FZ"} & named)
        if has_id and has_data and len(named) > len(set(best[1]) - {""}):
            best = (index, canonical)
    return best


def _units_from_row(row: Sequence[object], header: Sequence[str]) -> Optional[UnitSystem]:
    """Deduce el sistema de unidades del renglón de unidades de la tabla."""
    tokens = {h: _text(v) for h, v in zip(header, row) if _text(v)}
    if not tokens:
        return None
    length = tokens.get("station") or tokens.get("U3") or tokens.get("x") or ""
    force = tokens.get("V2") or tokens.get("P") or tokens.get("FZ") or ""
    if not force:
        moment = tokens.get("M3") or tokens.get("MZ") or ""
        force = moment.split("-")[0] if "-" in moment else ""
    if not (length or force):
        return None
    return parse_units_label(f"{force or 'tonf'} {length or 'm'} C")


def _is_units_row(row: Sequence[object], header: Sequence[str]) -> bool:
    """Un renglón de unidades no trae números en las columnas numéricas."""
    numeric_cols = [i for i, h in enumerate(header)
                    if h in ("station", "M3", "V2", "P", "U3", "drift", "FZ")]
    if not numeric_cols:
        return False
    for i in numeric_cols:
        if i < len(row) and _text(row[i]) and _num(row[i]) is not None:
            return False
    return any(i < len(row) and _text(row[i]) for i in numeric_cols)


def _cell(row: Sequence[object], header: Sequence[str], name: str) -> object:
    for i, h in enumerate(header):
        if h == name and i < len(row):
            return row[i]
    return None


def _has(header: Sequence[str], *names: str) -> bool:
    return any(n in header for n in names)


def parse_table_cells(cells: Sequence[Sequence[object]],
                      origin: str = "") -> ResultSet:
    """Interpreta una matriz de celdas ya leída."""
    result = ResultSet(source="tabla", origin=origin)
    blocks = _split_blocks(cells)
    parsed_any = False
    detected: List[str] = []

    for table_name, rows in blocks:
        header_index, header = _find_header(rows)
        if header_index < 0:
            continue
        body = list(rows[header_index + 1:])
        if body and _is_units_row(body[0], header):
            units = _units_from_row(body[0], header)
            if units is not None:
                result.units = units
            body = body[1:]

        kind = _classify(header)
        if kind is None:
            result.warnings.append(
                f"La tabla «{table_name or 'sin nombre'}» tiene columnas que no "
                "corresponden a fuerzas, desplazamientos, derivas ni reacciones.")
            continue
        detected.append(f"{table_name or kind}")
        _READERS[kind](body, header, result)
        parsed_any = True

    if not parsed_any:
        result.error = (
            "No se reconoció ninguna tabla de ETABS en el archivo. Exporta "
            "«Element Forces - Beams» (momentos y cortantes), «Joint "
            "Displacements» (deformaciones) o «Story Drifts» desde "
            "File ▸ Export ▸ Tables, en CSV o Excel.")
        return result

    result.origin = origin
    result.cases = sorted({r.case for r in result.frame_forces}
                          | {r.case for r in result.displacements}
                          | {r.case for r in result.drifts}
                          | {r.case for r in result.reactions})
    result.stories = sorted({r.story for r in result.frame_forces if r.story}
                            | {r.story for r in result.displacements if r.story}
                            | {r.story for r in result.drifts if r.story})
    if detected:
        result.warnings.insert(0, "Tablas leídas: " + ", ".join(detected) + ".")
    if not result.frame_forces:
        result.missing.append(
            "Elementos mecánicos por estación — falta la tabla «Element Forces»")
    if not result.displacements and not result.drifts:
        result.missing.append(
            "Deformaciones — faltan «Joint Displacements» o «Story Drifts»")
    if result.row_count == 0:
        result.error = "La tabla no trae renglones legibles."
    return result


def _classify(header: Sequence[str]) -> Optional[str]:
    if "station" in header and _has(header, "M3", "V2", "P", "M2", "V3", "T"):
        return "frame_forces"
    if _has(header, "U1", "U2", "U3", "R1", "R2", "R3") and "point" in header:
        return "displacements"
    if _has(header, "U1", "U2", "U3") and "story" in header:
        return "displacements"
    if "drift" in header:
        return "drifts"
    if _has(header, "FX", "FY", "FZ", "MX", "MY", "MZ"):
        return "reactions"
    return None


def _read_frame_forces(body: Sequence[Sequence[object]], header: Sequence[str],
                       result: ResultSet) -> None:
    if not _has(header, "element", "uniquename"):
        result.warnings.append(
            "La tabla de fuerzas no trae la columna Beam/Frame ni UniqueName: "
            "no se puede ligar cada renglón a una barra.")
        return
    if "V2" not in header:
        result.missing.append("Columna V2 (cortante) en la tabla de fuerzas")
    if "M3" not in header:
        result.missing.append("Columna M3 (momento) en la tabla de fuerzas")

    skipped = 0
    for row in body:
        element = _text(_cell(row, header, "element")) or \
            _text(_cell(row, header, "uniquename"))
        case = _text(_cell(row, header, "case"))
        station = _num(_cell(row, header, "station"))
        if not element or not case or station is None:
            if any(_text(c) for c in row):
                skipped += 1
            continue
        result.frame_forces.append(FrameForceRow(
            story=_text(_cell(row, header, "story")),
            frame=element,
            label=_text(_cell(row, header, "uniquename")) or None,
            case=case,
            step_type=_text(_cell(row, header, "steptype")) or None,
            step_num=_num(_cell(row, header, "stepnum")),
            station=station,
            P=_num(_cell(row, header, "P")),
            V2=_num(_cell(row, header, "V2")),
            V3=_num(_cell(row, header, "V3")),
            T=_num(_cell(row, header, "T")),
            M2=_num(_cell(row, header, "M2")),
            M3=_num(_cell(row, header, "M3")),
        ))
    if skipped:
        result.warnings.append(
            f"{skipped} renglón(es) de fuerzas se omitieron por venir "
            "incompletos (sin barra, caso o estación).")


def _read_displacements(body: Sequence[Sequence[object]], header: Sequence[str],
                        result: ResultSet) -> None:
    for row in body:
        case = _text(_cell(row, header, "case"))
        point = _text(_cell(row, header, "point")) or \
            _text(_cell(row, header, "uniquename"))
        values = {c: _num(_cell(row, header, c))
                  for c in ("U1", "U2", "U3", "R1", "R2", "R3")}
        if not case or all(v is None for v in values.values()):
            continue
        result.displacements.append(JointDisplRow(
            story=_text(_cell(row, header, "story")) or None,
            point=point or "(sin nudo)",
            label=_text(_cell(row, header, "label")) or None,
            case=case,
            step_type=_text(_cell(row, header, "steptype")) or None,
            step_num=_num(_cell(row, header, "stepnum")),
            **values))


def _read_drifts(body: Sequence[Sequence[object]], header: Sequence[str],
                 result: ResultSet) -> None:
    for row in body:
        case = _text(_cell(row, header, "case"))
        drift = _num(_cell(row, header, "drift"))
        if not case or drift is None:
            continue
        result.drifts.append(DriftRow(
            story=_text(_cell(row, header, "story")),
            case=case,
            step_type=_text(_cell(row, header, "steptype")) or None,
            direction=_text(_cell(row, header, "direction")) or None,
            drift=drift,
            label=_text(_cell(row, header, "label")) or None,
            x=_num(_cell(row, header, "x")),
            y=_num(_cell(row, header, "y")),
            z=_num(_cell(row, header, "z"))))


def _read_reactions(body: Sequence[Sequence[object]], header: Sequence[str],
                    result: ResultSet) -> None:
    for row in body:
        case = _text(_cell(row, header, "case"))
        values = {c: _num(_cell(row, header, c))
                  for c in ("FX", "FY", "FZ", "MX", "MY", "MZ")}
        if not case or all(v is None for v in values.values()):
            continue
        result.reactions.append(ReactionRow(
            case=case,
            step_type=_text(_cell(row, header, "steptype")) or None,
            **values))


_READERS = {
    "frame_forces": _read_frame_forces,
    "displacements": _read_displacements,
    "drifts": _read_drifts,
    "reactions": _read_reactions,
}


def parse_table_file(path: str) -> ResultSet:
    """Lee e interpreta una tabla exportada de ETABS."""
    cells = read_cells(path)
    return parse_table_cells(cells, os.path.basename(path))
