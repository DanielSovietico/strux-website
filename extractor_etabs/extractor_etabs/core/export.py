"""Exportadores: CSV, Excel y JSON.

Se exporta lo que se ve: las mismas columnas, en las mismas unidades y con el
encabezado de procedencia, para que un archivo suelto siga siendo trazable meses
después.
"""

from __future__ import annotations

import csv
import json
import os
from typing import Any, Dict, Iterable, List, Optional, Sequence

from .model import (
    TABLE_SPECS, ResultSet, StructuralModel, model_to_dict, table_columns,
    table_rows, to_dict,
)


def _provenance(result: ResultSet) -> List[List[Any]]:
    return [
        ["Extractor ETABS"],
        ["Fuente", result.source],
        ["Origen", result.origin or ""],
        ["Unidades", result.units.label],
        ["Casos", ", ".join(result.cases)],
        [],
    ]


def available_kinds(result: ResultSet) -> List[str]:
    """Tipos de tabla con al menos un renglón."""
    return [kind for kind in TABLE_SPECS
            if getattr(result, TABLE_SPECS[kind]["attr"], None)]


def export_csv(result: ResultSet, kind: str, path: str,
               with_header: bool = True) -> str:
    """Escribe una tabla en CSV (UTF-8 con BOM, que es lo que abre Excel bien)."""
    rows = table_rows(result, kind)
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        if with_header:
            writer.writerows(_provenance(result))
            writer.writerow([TABLE_SPECS[kind]["title"]])
        writer.writerow(table_columns(kind, result.units))
        writer.writerows(rows)
    return path


def export_csv_bundle(result: ResultSet, directory: str,
                      prefix: str = "extractor_etabs") -> List[str]:
    """Un CSV por tipo de tabla dentro de la carpeta indicada."""
    os.makedirs(directory, exist_ok=True)
    written: List[str] = []
    for kind in available_kinds(result):
        path = os.path.join(directory, f"{prefix}_{kind}.csv")
        written.append(export_csv(result, kind, path))
    return written


def export_xlsx(result: ResultSet, path: str,
                model: Optional[StructuralModel] = None) -> str:
    """Libro de Excel con una hoja por tabla, más una hoja de procedencia."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depende del entorno
        raise ImportError(
            "Para exportar a Excel se necesita openpyxl "
            "(pip install openpyxl). También puedes exportar a CSV.") from exc

    book = Workbook()
    info = book.active
    info.title = "Procedencia"
    info.append(["Extractor ETABS"])
    info["A1"].font = Font(bold=True, size=13)
    for row in _provenance(result)[1:]:
        info.append(row)
    if result.missing:
        info.append([])
        info.append(["Faltantes reportados"])
        for item in result.missing:
            info.append(["", item])
    if result.warnings:
        info.append([])
        info.append(["Avisos"])
        for item in result.warnings:
            info.append(["", item])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 110
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for kind in available_kinds(result):
        sheet = book.create_sheet(TABLE_SPECS[kind]["title"][:31])
        columns = table_columns(kind, result.units)
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in table_rows(result, kind):
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for index, name in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = \
                max(10, min(24, len(name) + 4))

    if model is not None:
        sheet = book.create_sheet("Barras del modelo")
        sheet.append(["Nivel", "Barra", "Etiqueta", "Tipo", "Sección",
                      f"Longitud ({model.info.units.length})", "Dirección",
                      "Eje", "De", "A"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for frame in model.frames:
            length = frame.length
            sheet.append([frame.story, frame.name, frame.label, frame.kind,
                          frame.section,
                          round(length, 4) if length is not None else None,
                          frame.direction, frame.grid_line, frame.grid_i,
                          frame.grid_j])
        sheet.freeze_panes = "A2"

    book.save(path)
    return path


def export_json(result: ResultSet, path: str,
                model: Optional[StructuralModel] = None) -> str:
    """Volcado completo en JSON (resultados y, opcionalmente, la geometría)."""
    payload: Dict[str, Any] = {"result": to_dict(result)}
    if model is not None:
        payload["model"] = model_to_dict(model)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, default=_default)
    return path


def export_model_json(model: StructuralModel, path: str) -> str:
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(model_to_dict(model), handle, ensure_ascii=False, indent=2,
                  default=_default)
    return path


def _default(value: Any) -> Any:      # pragma: no cover - red de seguridad
    return str(value)


def summary_text(result: ResultSet, model: Optional[StructuralModel] = None) -> str:
    """Resumen en texto plano, útil para el registro y para la CLI."""
    lines: List[str] = ["Extractor ETABS — resumen de extracción", ""]
    if model is not None:
        counts = ", ".join(f"{k}: {v}" for k, v in model.counts().items())
        lines += [f"Modelo: {model.info.file or '(sin nombre)'}",
                  f"Programa: {model.info.program or '?'} {model.info.version or ''}",
                  f"Unidades: {model.info.units.label}", f"Conteos: {counts}", ""]
    lines += [f"Fuente de resultados: {result.source}",
              f"Origen: {result.origin or '—'}",
              f"Unidades: {result.units.label}",
              f"Casos: {', '.join(result.cases) or '—'}"]
    for name, count in result.counts().items():
        if count:
            lines.append(f"  {name}: {count} renglones")
    if result.missing:
        lines += ["", "Faltantes:"] + [f"  · {m}" for m in result.missing]
    if result.warnings:
        lines += ["", "Avisos:"] + [f"  · {w}" for w in result.warnings]
    if result.error:
        lines += ["", f"Error: {result.error}"]
    return "\n".join(lines)
