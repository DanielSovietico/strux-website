"""Pruebas de la sesión, los exportadores y el puente con el visor original."""

from __future__ import annotations

import csv
import json
import os
import tempfile
import unittest

from extractor_etabs.core import export, strux_payload
from extractor_etabs.core.extractors import ExtractionOptions
from extractor_etabs.core.model import ResultSet, table_columns, table_rows
from extractor_etabs.core.session import Session, default_options

BASE = os.path.join(os.path.dirname(__file__), "..", "ejemplos")
E2K = os.path.join(BASE, "modelo_ejemplo.e2k")
TABLA = os.path.join(BASE, "element_forces_beams.csv")


def _sesion_completa() -> Session:
    session = Session()
    session.load_e2k(E2K)
    session.load_table(TABLA)
    session.add_deflections(default_options(session.model))
    return session


class TestSesion(unittest.TestCase):
    def test_flujo_sin_etabs(self) -> None:
        session = _sesion_completa()
        self.assertIsNotNone(session.model)
        self.assertIsNotNone(session.result)
        self.assertTrue(session.result.ok)
        self.assertGreater(len(session.result.deflections), 0)
        self.assertIn("archivo .e2k", session.source_label)
        self.assertIn("tabla exportada", session.source_label)
        self.assertFalse(session.can_extract)

    def test_deflexiones_sin_geometria_avisan(self) -> None:
        session = Session()
        session.load_table(TABLA)
        with self.assertRaises(Exception) as ctx:
            session.add_deflections(ExtractionOptions())
        self.assertIn("geometría", str(ctx.exception))

    def test_deflexiones_no_se_duplican_al_recalcular(self) -> None:
        session = _sesion_completa()
        primero = len(session.result.deflections)
        session.add_deflections(default_options(session.model))
        self.assertEqual(len(session.result.deflections), primero)

    def test_factor_de_inercia_aumenta_la_flecha(self) -> None:
        session = _sesion_completa()
        peor = max(abs(d.deflection or 0.0) for d in session.result.deflections)
        options = default_options(session.model)
        options.inertia_factor = 0.5
        session.add_deflections(options)
        agrietada = max(abs(d.deflection or 0.0)
                        for d in session.result.deflections)
        self.assertAlmostEqual(agrietada, peor * 2.0, delta=peor * 0.02)

    def test_registro_acumula_mensajes(self) -> None:
        mensajes = []
        session = Session(on_log=mensajes.append)
        session.load_e2k(E2K)
        self.assertTrue(mensajes)
        self.assertEqual(mensajes, session.log)


class TestTablasPlanas(unittest.TestCase):
    def test_encabezados_llevan_unidades(self) -> None:
        session = _sesion_completa()
        columnas = table_columns("frame_forces", session.result.units)
        self.assertIn("M3 (tonf-m)", columnas)
        self.assertIn("Estación (m)", columnas)
        self.assertIn("V2 (tonf)", columnas)

    def test_renglones_coinciden_con_las_columnas(self) -> None:
        session = _sesion_completa()
        for kind in ("frame_forces", "deflections"):
            columnas = table_columns(kind, session.result.units)
            renglones = table_rows(session.result, kind)
            self.assertTrue(renglones)
            self.assertTrue(all(len(r) == len(columnas) for r in renglones))


class TestExportadores(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.session = _sesion_completa()

    def test_csv_incluye_procedencia_y_datos(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = export.export_csv(self.session.result, "frame_forces",
                                     os.path.join(folder, "f.csv"))
            with open(path, encoding="utf-8-sig", newline="") as handle:
                filas = list(csv.reader(handle))
        self.assertEqual(filas[0][0], "Extractor ETABS")
        self.assertIn("element_forces_beams.csv", filas[2][1])
        encabezado = next(f for f in filas if f and f[0] == "Nivel")
        self.assertIn("M3 (tonf-m)", encabezado)
        self.assertEqual(len(filas) - filas.index(encabezado) - 1, 4284)

    def test_paquete_de_csv(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            rutas = export.export_csv_bundle(self.session.result, folder)
        nombres = sorted(os.path.basename(r) for r in rutas)
        self.assertEqual(nombres, ["extractor_etabs_deflections.csv",
                                   "extractor_etabs_frame_forces.csv"])

    def test_json_trae_modelo_y_resultados(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = export.export_json(self.session.result,
                                      os.path.join(folder, "t.json"),
                                      self.session.model)
            with open(path, encoding="utf-8") as handle:
                data = json.load(handle)
        self.assertIn("frame_forces", data["result"]["tables"])
        self.assertEqual(data["result"]["units"]["length"], "m")
        self.assertEqual(len(data["model"]["frames"]), 116)
        self.assertEqual(data["model"]["counts"]["trabes"], 68)

    def test_excel(self) -> None:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl no está instalado")
        with tempfile.TemporaryDirectory() as folder:
            path = export.export_xlsx(self.session.result,
                                      os.path.join(folder, "t.xlsx"),
                                      self.session.model)
            book = openpyxl.load_workbook(path, read_only=True)
            hojas = book.sheetnames
            fuerzas = book["Elementos mecánicos (barras)"]
            filas = fuerzas.max_row
            book.close()
        self.assertIn("Procedencia", hojas)
        self.assertIn("Barras del modelo", hojas)
        self.assertEqual(filas, 4285)          # encabezado + renglones

    def test_resumen_menciona_faltantes(self) -> None:
        texto = export.summary_text(self.session.result, self.session.model)
        self.assertIn("Extractor ETABS", texto)
        self.assertIn("Faltantes:", texto)
        self.assertIn("fuerzas: 4284", texto)

    def test_resumen_de_resultado_vacio(self) -> None:
        texto = export.summary_text(ResultSet(source="—"))
        self.assertIn("Fuente de resultados: —", texto)


class TestPuenteVisor(unittest.TestCase):
    def test_forma_esperada_por_el_visor(self) -> None:
        session = _sesion_completa()
        with tempfile.TemporaryDirectory() as folder:
            path = strux_payload.write_bridge_json(
                os.path.join(folder, "p.json"), session.model, session.result)
            with open(path, encoding="utf-8") as handle:
                data = json.load(handle)

        modelo = data["etabs_model"]
        self.assertEqual(modelo["program"], "ETABS")
        self.assertEqual(len(modelo["beams"]), 68)
        self.assertEqual(sorted(modelo["beams"][0]),
                         ["id", "section", "story", "xi", "xj", "yi", "yj"])
        self.assertEqual(modelo["gx"][0], ["A", 0])
        self.assertIn("V-30X70", modelo["sections"])
        self.assertEqual(modelo["sections"]["V-30X70"]["h"], 70.0)

        fuerzas = data["etabs_frame_forces"]["frames"]
        self.assertEqual(len(fuerzas), 68)
        primero = fuerzas[0]
        self.assertEqual(sorted(primero),
                         ["loadCase", "m3", "name", "p", "station", "story",
                          "v2"])
        self.assertEqual(len(primero["station"]), len(primero["m3"]))
        self.assertEqual(len(primero["station"]), len(primero["loadCase"]))

    def test_filtra_por_nivel_y_caso(self) -> None:
        session = _sesion_completa()
        payload = strux_payload.frame_forces_payload(
            session.result, story="Azotea", cases=["1.3CM+1.5CV"])
        self.assertEqual(len(payload["frames"]), 17)
        for frame in payload["frames"]:
            self.assertEqual(frame["story"], "Azotea")
            self.assertEqual(set(frame["loadCase"]), {"1.3CM+1.5CV"})


if __name__ == "__main__":
    unittest.main()
